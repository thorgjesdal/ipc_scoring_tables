from openpyxl import load_workbook
from ipc_score import ipc_score

fname = 'NM_para_2024.xlsx'
wb = load_workbook(filename=fname)
ws = wb.active
#for value in ws.iter_rows(min_row=2,min_col=1, max_col=6, values_only=True):
for i, value in enumerate( ws.iter_rows(min_row=2,min_col=1, max_col=6, values_only=True) ):
    if value[0] is not None:
        name = value[0]
        event = f'{value[1]}'
        gender = value[2]
        cat = value[3]
        perf = value[4]
        print(perf)
        if perf is not None :
            perf = f'{perf}'
            points = ipc_score(event, gender, cat, perf, custom='NOR')
            print(i, name, event, gender, perf, points)
            """
            if points < 100:
                points = 0.5*(points+100)
                """
            ws[f'F{i+2}'] = points

outfile = fname.replace('.xlsx', '_pts.xlsx')
wb.save(outfile)

