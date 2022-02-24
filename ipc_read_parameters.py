import sys
import pprint
import json
from openpyxl import load_workbook

gender = { 'Men' : 'M', 'M' : 'M', 'Women' : 'F', 'W' : 'F'}

# opentrack/standard event codes. use 'OT' for Club Throw until it gets its own code
event_codes = { '100 m' : '100', '200 m' : '200', '400 m': '400', 
                '800 m' :'800', '1500 m' : '1500', '5000 m' : '5000', '10000 m':'10000', 
                'Shot Put' : 'SP', 'Discus' : 'DT', 'Javelin' : 'JT', 'Club Throw' : 'OT',
                'High Jump' : 'HJ', 'Long Jump' : 'LJ', 'Triple Jump' : 'TJ' }


ipc_parameters = {}
def ipc_read_parameters(f):
    wb = load_workbook(filename=f)
    ws = wb["Parameters"]
    params = {}
    for value in ws.iter_rows(min_row=2,min_col=1, max_col=6, values_only=True):
        print(value)
        e = event_codes[ value[1] ]
        g = gender[ value[0] ]
        c = value[2]
        if e not in ipc_parameters.keys():
            ipc_parameters[e] = {}
        if g not in ipc_parameters[e].keys():
            ipc_parameters[e][g] = {}
        if c not in ipc_parameters[e][g].keys():
            ipc_parameters[e][g][c] = value[3:6]

def save_as_python(p):
    if isinstance(p, dict):
        with open('ipc_parameters.pyi', 'w') as f:
            pp = pprint.PrettyPrinter(indent=4, stream=f)
            pp.pprint(p)
   
def save_as_json(p):
    if isinstance(p, dict):
        with open('ipc_parameters.json', 'w') as f:
            json.dump(p, f)


   

for i in range(1,len(sys.argv)):
    infile = sys.argv[i]
    ipc_read_parameters(infile)

save_as_python(ipc_parameters)       
save_as_json(ipc_parameters)       
